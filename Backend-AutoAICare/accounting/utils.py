"""
Utility functions for generating PDF and Excel reports for accounting module
"""
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime


def generate_profit_loss_pdf(data, start_date, end_date, branch_name=None):
    """Generate Profit & Loss statement PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("Profit & Loss Statement", title_style)
    elements.append(title)
    
    # Period info
    period_text = f"Period: {start_date} to {end_date}"
    if branch_name:
        period_text += f" | Branch: {branch_name}"
    period = Paragraph(period_text, styles['Normal'])
    elements.append(period)
    elements.append(Spacer(1, 20))
    
    # Income table
    income_data = [['Income', 'Amount']]
    for item in data.get('income', []):
        income_data.append([item['description'], f"₹{item['amount']:,.2f}"])
    income_data.append(['Total Income', f"₹{data.get('total_income', 0):,.2f}"])
    
    income_table = Table(income_data, colWidths=[4*inch, 2*inch])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#22c55e')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(income_table)
    elements.append(Spacer(1, 20))
    
    # Expense table
    expense_data = [['Expenses', 'Amount']]
    for item in data.get('expenses', []):
        expense_data.append([item.get('category_display', item.get('description', '')), f"₹{item['amount']:,.2f}"])
    expense_data.append(['Total Expenses', f"₹{data.get('total_expenses', 0):,.2f}"])
    
    expense_table = Table(expense_data, colWidths=[4*inch, 2*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(expense_table)
    elements.append(Spacer(1, 20))
    
    # Net Profit
    net_profit = data.get('net_profit', 0)
    profit_color = colors.HexColor('#22c55e') if net_profit >= 0 else colors.HexColor('#ef4444')
    profit_data = [['Net Profit/Loss', f"₹{net_profit:,.2f}"]]
    profit_table = Table(profit_data, colWidths=[4*inch, 2*inch])
    profit_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), profit_color),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('PADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(profit_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_profit_loss_excel(data, start_date, end_date, branch_name=None):
    """Generate Profit & Loss statement Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    # Styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Period
    ws.merge_cells('A2:B2')
    period_text = f"Period: {start_date} to {end_date}"
    if branch_name:
        period_text += f" | Branch: {branch_name}"
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    
    # Income section
    ws[f'A{row}'] = "Income"
    ws[f'B{row}'] = "Amount"
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    for item in data.get('income', []):
        ws[f'A{row}'] = item['description']
        ws[f'B{row}'] = item['amount']
        ws[f'B{row}'].number_format = '₹#,##0.00'
        row += 1
    
    ws[f'A{row}'] = "Total Income"
    ws[f'B{row}'] = data.get('total_income', 0)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '₹#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 2
    
    # Expense section
    ws[f'A{row}'] = "Expenses"
    ws[f'B{row}'] = "Amount"
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    row += 1
    
    for item in data.get('expenses', []):
        ws[f'A{row}'] = item.get('category_display', item.get('description', ''))
        ws[f'B{row}'] = item['amount']
        ws[f'B{row}'].number_format = '₹#,##0.00'
        row += 1
    
    ws[f'A{row}'] = "Total Expenses"
    ws[f'B{row}'] = data.get('total_expenses', 0)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'B{row}'].number_format = '₹#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    row += 2
    
    # Net Profit
    net_profit = data.get('net_profit', 0)
    profit_color = "22C55E" if net_profit >= 0 else "EF4444"
    ws[f'A{row}'] = "Net Profit/Loss"
    ws[f'B{row}'] = net_profit
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'B{row}'].number_format = '₹#,##0.00'
    ws[f'B{row}'].fill = PatternFill(start_color=profit_color, end_color=profit_color, fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_expense_report_excel(expenses, start_date, end_date, branch_name=None):
    """Generate Expense Report Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Expense Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Period
    ws.merge_cells('A2:D2')
    period_text = f"Period: {start_date} to {end_date}"
    if branch_name:
        period_text += f" | Branch: {branch_name}"
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Date', 'Category', 'Title', 'Amount']
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 5
    total = 0
    for expense in expenses:
        ws.cell(row=row, column=1, value=expense.get('date', ''))
        ws.cell(row=row, column=2, value=expense.get('category_display', ''))
        ws.cell(row=row, column=3, value=expense.get('title', ''))
        amount_cell = ws.cell(row=row, column=4, value=expense.get('amount', 0))
        amount_cell.number_format = '₹#,##0.00'
        total += expense.get('amount', 0)
        row += 1
    
    # Total
    ws.cell(row=row, column=3, value="Total:").font = Font(bold=True)
    total_cell = ws.cell(row=row, column=4, value=total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '₹#,##0.00'
    total_cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    total_cell.font = Font(bold=True, color="FFFFFF")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
