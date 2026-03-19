"""
Advanced Reporting Utilities
Handles PDF and Excel report generation
"""
from io import BytesIO
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Avg, Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from bookings.models import Booking


class PDFReportGenerator:
    """Generate PDF reports with professional formatting"""
    
    def __init__(self, title="Business Report", company_name="DetailEase"):
        self.title = title
        self.company_name = company_name
        self.buffer = BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        self.elements = []
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER,
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['BodyText'],
            fontSize=10,
            spaceAfter=12,
        ))
    
    def add_header(self):
        """Add report header"""
        # Company name
        company = Paragraph(self.company_name, self.styles['CustomTitle'])
        self.elements.append(company)
        
        # Report title
        title = Paragraph(self.title, self.styles['CustomHeading'])
        self.elements.append(title)
        
        # Date
        date_str = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        date = Paragraph(date_str, self.styles['CustomBody'])
        self.elements.append(date)
        
        self.elements.append(Spacer(1, 20))
    
    def add_section(self, title, content=None):
        """Add a section with title and optional content"""
        section_title = Paragraph(title, self.styles['CustomHeading'])
        self.elements.append(section_title)
        
        if content:
            section_content = Paragraph(content, self.styles['CustomBody'])
            self.elements.append(section_content)
        
        self.elements.append(Spacer(1, 12))
    
    def add_table(self, data, col_widths=None, style_name='default'):
        """Add a formatted table"""
        if not data:
            return
        
        # Create table
        table = Table(data, colWidths=col_widths)
        
        # Apply style
        if style_name == 'default':
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ])
        else:
            table_style = TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ])
        
        table.setStyle(table_style)
        self.elements.append(table)
        self.elements.append(Spacer(1, 20))
    
    def add_summary_box(self, title, value, description=""):
        """Add a summary statistics box"""
        data = [
            [str(title)],
            [str(value)]
        ]
        if description:
            data.append([str(description)])
        
        table = Table(data, colWidths=[4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, 1), 24),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1e40af')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        self.elements.append(table)
        self.elements.append(Spacer(1, 12))
    
    def generate(self):
        """Generate the PDF and return buffer"""
        self.doc.build(self.elements)
        self.buffer.seek(0)
        return self.buffer


class ExcelReportGenerator:
    """Generate Excel reports with professional formatting"""
    
    def __init__(self, title="Business Report"):
        self.title = title
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.current_sheet = None
        
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        self.title_font = Font(bold=True, size=16, color="1e40af")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_sheet(self, name):
        """Add a new worksheet"""
        self.current_sheet = self.workbook.create_sheet(title=name)
        return self.current_sheet
    
    def add_title(self, title, row=1):
        """Add title to current sheet"""
        if not self.current_sheet:
            return
        
        cell = self.current_sheet.cell(row=row, column=1)
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center')
        
        # Merge cells for title
        self.current_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    
    def add_header_row(self, headers, row=3):
        """Add header row with formatting"""
        if not self.current_sheet:
            return
        
        for col, header in enumerate(headers, start=1):
            cell = self.current_sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def add_data_rows(self, data, start_row=4):
        """Add data rows"""
        if not self.current_sheet:
            return
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.current_sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
    
    def auto_adjust_columns(self):
        """Auto-adjust column widths"""
        if not self.current_sheet:
            return
        
        for column in self.current_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.current_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_summary_section(self, summaries, start_row=1):
        """Add summary statistics section"""
        if not self.current_sheet:
            return
        
        current_row = start_row
        for label, value in summaries.items():
            label_cell = self.current_sheet.cell(row=current_row, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            
            value_cell = self.current_sheet.cell(row=current_row, column=2)
            value_cell.value = value
            
            current_row += 1
        
        return current_row + 1  # Return next available row
    
    def generate(self):
        """Generate Excel file and return buffer"""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


def generate_revenue_report_pdf(start_date, end_date, bookings_data):
    """Generate revenue report in PDF format"""
    
    # Create PDF generator
    pdf = PDFReportGenerator(
        title=f"Revenue Report ({start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')})"
    )
    
    # Add header
    pdf.add_header()
    
    # Calculate summary statistics
    total_revenue = bookings_data.aggregate(total=Sum('total_price'))['total'] or 0
    total_bookings = bookings_data.count()
    avg_booking_value = bookings_data.aggregate(avg=Avg('total_price'))['avg'] or 0
    
    # Add summary boxes
    pdf.add_section("Summary Statistics")
    pdf.add_summary_box("Total Revenue", f"₹{total_revenue:,.2f}")
    pdf.add_summary_box("Total Bookings", total_bookings)
    pdf.add_summary_box("Average Booking Value", f"₹{avg_booking_value:,.2f}")
    
    # Add detailed table
    pdf.add_section("Booking Details")
    
    table_data = [['Date', 'Customer', 'Service', 'Amount', 'Status']]
    for booking in bookings_data[:50]:  # Limit to 50 for PDF
        table_data.append([
            booking.booking_datetime.strftime('%Y-%m-%d'),
            booking.customer.user.name if booking.customer else 'N/A',
            booking.package.name if booking.package else 'N/A',
            f"₹{booking.total_price:,.2f}",
            booking.status
        ])
    
    pdf.add_table(table_data)
    
    return pdf.generate()


def generate_revenue_report_excel(start_date, end_date, bookings_data):
    """Generate revenue report in Excel format"""
    excel = ExcelReportGenerator(title="Revenue Report")
    
    # Create main sheet
    sheet = excel.add_sheet("Revenue Report")
    
    # Add title
    excel.add_title(f"Revenue Report ({start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')})")
    
    # Add summary section
    total_revenue = bookings_data.aggregate(total=Sum('total_price'))['total'] or 0
    total_bookings = bookings_data.count()
    avg_booking_value = bookings_data.aggregate(avg=Avg('total_price'))['avg'] or 0
    
    summaries = {
        'Total Revenue': f"₹{total_revenue:,.2f}",
        'Total Bookings': total_bookings,
        'Average Booking Value': f"₹{avg_booking_value:,.2f}",
        'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    next_row = excel.add_summary_section(summaries, start_row=3)
    
    # Add data table
    headers = ['Date', 'Customer', 'Service', 'Amount', 'Status', 'Payment Status']
    excel.add_header_row(headers, row=next_row + 1)
    
    data_rows = []
    for booking in bookings_data:
        data_rows.append([
            booking.booking_datetime.strftime('%Y-%m-%d'),
            booking.customer.user.name if booking.customer else 'N/A',
            booking.package.name if booking.package else 'N/A',
            booking.total_price,
            booking.status,
            'Paid' if booking.payments.filter(payment_status='completed').exists() else 'Pending'
        ])
    
    excel.add_data_rows(data_rows, start_row=next_row + 2)
    excel.auto_adjust_columns()
    
    return excel.generate()


def generate_customer_report_pdf(customers_data):
    """Generate customer report in PDF format"""
    # Create PDF generator
    pdf = PDFReportGenerator(title="Customer Report")
    pdf.add_header()
    
    # Summary - using pre-filtered queryset for efficiency
    total_customers = customers_data.count()
    active_customers = customers_data.filter(user__is_active=True).count()
    
    # Calculate total revenue across all these customers
    total_revenue_stats = Booking.objects.filter(customer__in=customers_data).aggregate(
        total=Sum('total_price')
    )
    total_revenue = total_revenue_stats['total'] or 0
    
    pdf.add_section("Customer Statistics")
    pdf.add_summary_box("Total Customers", total_customers)
    pdf.add_summary_box("Active Customers", active_customers)
    pdf.add_summary_box("Total Revenue", f"₹{float(total_revenue):,.2f}")
    
    # Customer table
    pdf.add_section("Customer Details")
    table_data = [['Name', 'Phone', 'Email', 'Revenue', 'Status']]
    
    # Annotate customers with their total revenue for efficiency (N+1 fix)
    annotated_customers = customers_data.annotate(
        total_spent=Sum('bookings__total_price')
    ).order_by('-total_spent')[:50]
    
    for customer in annotated_customers:
        spent = customer.total_spent or 0
        table_data.append([
            customer.user.name,
            customer.user.phone,
            customer.user.email or 'N/A',
            f"₹{float(spent):,.2f}",
            'Active' if customer.user.is_active else 'Inactive'
        ])
    
    pdf.add_table(table_data)
    
    return pdf.generate()


def generate_customer_report_excel(customers_data):
    """Generate customer report in Excel format"""
    excel = ExcelReportGenerator(title="Customer Report")
    sheet = excel.add_sheet("Customers")
    
    excel.add_title("Customer Report")
    
    # Summary
    total_customers = customers_data.count()
    active_customers = customers_data.filter(user__is_active=True).count()
    
    # Calculate total revenue
    total_revenue_stats = Booking.objects.filter(customer__in=customers_data).aggregate(
        total=Sum('total_price')
    )
    total_revenue = total_revenue_stats['total'] or 0
    
    summaries = {
        'Total Customers': total_customers,
        'Active Customers': active_customers,
        'Total Revenue': f"₹{float(total_revenue):,.2f}",
        'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    next_row = excel.add_summary_section(summaries, start_row=3)
    
    # Data table
    headers = ['Name', 'Phone', 'Email', 'Revenue', 'Status']
    excel.add_header_row(headers, row=next_row + 1)
    
    # Annotate customers with their total revenue for efficiency (N+1 fix)
    annotated_customers = customers_data.annotate(
        total_spent=Sum('bookings__total_price')
    ).order_by('-total_spent')
    
    data_rows = []
    for customer in annotated_customers:
        spent = customer.total_spent or 0
        data_rows.append([
            customer.user.name,
            customer.user.phone,
            customer.user.email or 'N/A',
            float(spent),
            'Active' if customer.user.is_active else 'Inactive'
        ])
    
    excel.add_data_rows(data_rows, start_row=next_row + 2)
    excel.auto_adjust_columns()
    
    return excel.generate()
